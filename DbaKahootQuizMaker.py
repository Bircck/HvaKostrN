import os
import sys
import openpyxl as xl
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from urllib.parse import urlparse, urljoin
from copy import copy
from fake_useragent import UserAgent
import re
import random
import urllib.request
import numpy as np
import time
import sys
import argparse

# # named command line arguments
# parser=argparse.ArgumentParser()
# parser.add_argument("--amount", help="amount of questions below 35")
# parser.add_argument("--name", help="name of the quiz")
# parser.add_argument("--openbrowser", help="whether to open a window with the quiz or not, default false")

# args=parser.parse_args()

class DBAlisting:
    def __init__(self, name, price, imageurl):
        self.name = name
        self.price = price
        self.imageurl = imageurl

def CreateDirectory(dirName):
    """Creates a new quiz directory and returns path
    """
    # Parent Directory path
    currentWorkFolder = os.getcwd()
    ExportDirectory = 'ExportedQuizzes'
    parentDirPath = os.path.join(currentWorkFolder, ExportDirectory)

    if (not os.path.exists(parentDirPath)):
        os.mkdir(parentDirPath)
    
    # Path
    path = os.path.join(parentDirPath, dirName)
    
    # Create the directory
    if (not os.path.exists(path)):
        os.mkdir(path)
        print("Directory '% s' created" % path)
    else:
        print("A quiz with this name already exists, Please delete the folder or choose another name")
        quit()
    return path

def GeneratePrices(price):
    price = int(price)
    lowerPricesAmount = random.randint(0,3)

    multipliersLower = []
    multipliersHigher = []
    # multipliers = []
    multipliersLower.extend(np.arange(0.05,1,0.05))
    multipliersLower.extend(np.arange(0.2, 1, 0.2))
    multipliersHigher.extend(np.arange(2,21))
    multipliersHigher.extend(np.arange(1.5,5,0.5))
    multipliersHigher.extend(np.arange(1.2, 2.2, 0.2))

    results = []
    while len(results) != lowerPricesAmount:
        newPrice = int(price*multipliersLower[random.randint(0,len(multipliersLower)-1)])
        if not (newPrice in results):
            results.append(newPrice)

    while len(results) < 3:
        newPrice = int(price*multipliersHigher[random.randint(0,len(multipliersHigher)-1)])
        if not (newPrice in results):
            results.append(newPrice)
    
    results.append(price)

    random.shuffle(results)
    return results

def CreateKahootExcel(dirPath, name, listings):
    # region kahoot template values
    # startrow, start-column and end column for kahoot template
    startR = 9
    startC = 2
    endC = 8
    
    # specific columns
    qC = 2  # question column
    a1C = 3 # answer 1 column
    a2C = 4 # answer 2 column
    a3C = 5 # answer 3 column
    a4C = 6 # answer 4 column
    tC = 7  # time limit column
    cqC = 8 # correct questions column (comma seperated list ex: 1,2,4)
    # endregion

    # opening the source excel file
    templateFilename ="KahootQuizTemplate.xlsx"
    wb1 = xl.load_workbook(templateFilename)
    ws1 = wb1.worksheets[0]
    excelExportPath = f"{dirPath}"+"\\"+name+".xlsx"

    # set values and get images 
    qIndex = 0 
    for i in range (startR, startR+len(listings)):
        listing:DBAlisting = listings[qIndex]
        title = listing.name
        price = listing.price
        intPrice = int(re.sub("[^0-9]", "", listing.price))
        prices = GeneratePrices(intPrice)
        correctPriceIndex = prices.index(intPrice)

        # insert values
        ws1.cell(row = i, column = qC).value = f"Hvad koster {title[0:100]}?"
        ws1.cell(row = i, column = a1C).value = f"{prices[0]} kr."
        ws1.cell(row = i, column = a2C).value = f"{prices[1]} kr."
        ws1.cell(row = i, column = a3C).value = f"{prices[2]} kr."
        ws1.cell(row = i, column = a4C).value = f"{prices[3]} kr."
        ws1.cell(row = i, column = tC).value = 20
        ws1.cell(row = i, column = cqC).value = correctPriceIndex+1
        
        # save image
        urllib.request.urlretrieve(listing.imageurl, f"{dirPath}\\question{qIndex+1}.jpg")
        qIndex +=1

    #save
    if (not os.path.exists(excelExportPath)):
        wb1.save(excelExportPath)    
        print(f"Kahoot Created at {dirPath}")
    else: 
        print("file already exists and hasn't been overwritten")

    return

def GetDbaData(url, entries = 5):
    # Make a request to the website and get the HTML content
    ua = UserAgent()
    headers = {'User-Agent': ua.random}

    trListings = []

    if entries > 20:
        response2 = requests.get("https://www.dba.dk/soeg/?sort=listingdate&per=3", headers=headers)
        soup2 = BeautifulSoup(response2.content, 'html.parser')
        trListings += soup2.select('tr.dbaListing.listing', limit = entries - 20)
        entries = 20

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    trListings += soup.select('tr.dbaListing.listing', limit = entries)
    dbaListings = []
    for tr in trListings: 
        #title
        mainContent = tr.find("td", class_="mainContent")
        name = mainContent.find("span", class_="text").text
        price = mainContent.find("span", class_="price").text
        #price
        # listing = BeautifulSoup(str(tr), 'html.parser')
        # tds = listing.find_all("td")
        # pricetd = tds[3]
        # price = pricetd.text
        #img
        # div = tr.find("div", class_="thumbnailContainer")
        img = tr.find("img", class_="image-thumbnail")
        imgSRC = img['src']

        # url = div['data-original']
        # image = url.replace('url(', '').replace(')', '') 
        # image = url.replace('?class=S200X200', '') 

        dbalisting = DBAlisting(name, price, imgSRC)
        dbaListings.append(dbalisting)

    return dbaListings

# url og query for random listings https://www.dba.dk/soeg/?sort=listingdate&per=2
# listings object = {title: "abc", price: "123"}

def QuestionAmountAndTypeIsCorrect(inputAmount):
    if inputAmount.isnumeric():
        questionsAmount = int(inputAmount)
        if questionsAmount < 36:
            return True
    return False

def CreateQuiz(questionsAmount = "", quizName = None):
    #define name and questions amount
    if not QuestionAmountAndTypeIsCorrect(questionsAmount):
        while not isinstance(questionsAmount, int):
            print('Enter Amount of questions (below 35): ')
            inputAmount = input()
            if QuestionAmountAndTypeIsCorrect(inputAmount):
                questionsAmount = int(inputAmount)        
    else:
        questionsAmount = int(questionsAmount)

    if quizName == None:
        quizName = "hvadkoster"+str(random.randint(0,10000))

    #create quiz
    exportDirPath = CreateDirectory(quizName)

    url = "https://www.dba.dk/soeg/?sort=listingdate&per=2"
    listings = GetDbaData(url, questionsAmount)
    CreateKahootExcel(exportDirPath, quizName, listings)
    return exportDirPath    

if __name__== "__main__":
    if len(sys.argv) == 3:
        CreateQuiz(sys.argv[1], sys.argv[2])
    else:
        CreateQuiz()
    # CreateQuiz(args.amount, args.name,)
